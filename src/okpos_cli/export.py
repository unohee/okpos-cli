"""xlsx export.

One worksheet per controller. The JSONB payload is flattened back into columns
using the union of keys seen in that controller's rows, so the sheet looks like
the original OKPOS grid rather than a JSON blob.
"""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_INVALID_SHEET = re.compile(r"[\[\]:*?/\\]")
META_COLUMNS = ["biz_date", "shop_cd", "sheet_seq", "row_no", "scraped_at"]


def _sheet_title(name: str, used: set[str]) -> str:
    title = _INVALID_SHEET.sub("_", name)[:31] or "sheet"
    base, n = title, 1
    while title in used:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def write_xlsx(records: list[dict[str, Any]], out_path: Path) -> tuple[Path, int, int]:
    """Write records to xlsx, grouped per controller. Returns (path, sheets, rows)."""
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for rec in records:
        grouped[rec["controller"]].append(rec)

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4C5667")
    used: set[str] = set()
    total = 0

    for controller, rows in sorted(grouped.items()):
        payload_keys: list[str] = []
        for rec in rows:
            for k in (rec["payload"] or {}):
                if k not in payload_keys:
                    payload_keys.append(k)

        ws = wb.create_sheet(_sheet_title(controller.split(".")[-1], used))
        ws.append(META_COLUMNS + payload_keys)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for rec in rows:
            payload = rec["payload"] or {}
            scraped = rec.get("scraped_at")
            ws.append(
                [
                    rec["biz_date"],
                    rec["shop_cd"],
                    rec["sheet_seq"],
                    rec["row_no"],
                    scraped.replace(tzinfo=None) if isinstance(scraped, datetime) else scraped,
                ]
                + [payload.get(k) for k in payload_keys]
            )
            total += 1

        ws.freeze_panes = "A2"
        for idx in range(1, len(META_COLUMNS) + len(payload_keys) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 16

    if not grouped:
        wb.create_sheet("empty").append(["no records matched the filter"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, len(grouped), total


def default_export_name(date_from: date | None, date_to: date | None) -> str:
    span = ""
    if date_from and date_to:
        span = f"_{date_from.isoformat()}_{date_to.isoformat()}"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"okpos{span}_{stamp}.xlsx"
